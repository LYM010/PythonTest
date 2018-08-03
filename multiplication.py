#!/usr/bin/env python3
# coding:UTF-8

import logging,os,openpyxl,sys
from openpyxl.styles import Font

logging.basicConfig(level=logging.DEBUG,format='%(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.CRITICAL)

os.chdir(os.path.join('..','tmp'))
num = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb['Sheet']
logging.debug(f'sheet is {sheet}')

logging.debug('Setting Font Style...')
fontObj = Font(bold = True)

logging.debug('Setting title...')
for i in range(1,num+2):
	if i == 1:
		sheet.cell(row=1,column=1).value = ''
		sheet.cell(row=1,column=1).font=fontObj
		continue
	sheet.cell(row=i,column=1).value = i-1
	sheet.cell(row=i,column=1).font=fontObj
	sheet.cell(row=1,column=i).value = i-1
	sheet.cell(row=1,column=i).font=fontObj
logging.debug('inserting data...')
for i in range(2,num+2):
	for j in range(2,num+2):
		sheet.cell(row=i,column=j).value=(i-1)*(j-1)

logging.debug('Saving File...')
wb.save('multiplicationTable.xlsx')