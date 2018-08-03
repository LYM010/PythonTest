#!/usr/bin/env python3
# coding:UTF-8
import openpyxl,logging,os
logging.basicConfig(level=logging.DEBUG,format='%(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.CRITICAL)

os.chdir(os.path.join('..','tmp'))
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']
logging.debug(f'worksheet is {sheet}')

PRICE_UPDATES = {'Garlic':3.07,'Celery':1.19,'Lemon':1.27}

for rowNum in range(2,sheet.max_row):
    produceName = sheet.cell(row=rowNum,column=1).value
    logging.debug(f'produceName is {produceName}')
    if produceName in PRICE_UPDATES:
        sheet.cell(row = rowNum,column=2).value = PRICE_UPDATES[produceName]
        logging.debug(f'sheet.cell(row=rowNum,column=2')

wb.save('updateProduceSales.xlsx')
logging.debug(f'Program Done.')
